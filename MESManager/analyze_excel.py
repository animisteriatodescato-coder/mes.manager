import openpyxl
from openpyxl.utils import get_column_letter
import re

wb = openpyxl.load_workbook(r'c:\Dev\MESManager\TABELLA PREZZI MODIFICATA.xlsx', data_only=False)

print('=' * 80)
print('ANALISI FILE EXCEL: TABELLA PREZZI MODIFICATA.xlsx')
print('=' * 80)

# 1. Lista fogli
print('\n## FOGLI PRESENTI:')
for i, sheet_name in enumerate(wb.sheetnames, 1):
    ws = wb[sheet_name]
    print(f'  {i}. {sheet_name} (righe: {ws.max_row}, colonne: {ws.max_column})')

# 2. Analisi formule per foglio
print('\n## ANALISI FORMULE PER FOGLIO:')
all_formulas = []
cross_sheet_refs = []
formula_cells = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    sheet_formulas = []
    
    for row in range(1, min(ws.max_row + 1, 500)):
        for col in range(1, min(ws.max_column + 1, 50)):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                cell_ref = f'{get_column_letter(col)}{row}'
                sheet_formulas.append({
                    'cell': cell_ref,
                    'formula': formula
                })
                
                # Cerca riferimenti ad altri fogli
                if "'" in formula or '!' in formula:
                    cross_sheet_refs.append({
                        'sheet': sheet_name,
                        'cell': cell_ref,
                        'formula': formula
                    })
    
    if sheet_formulas:
        formula_cells[sheet_name] = sheet_formulas
        print(f'\n### Foglio: {sheet_name}')
        print(f'    Totale formule trovate: {len(sheet_formulas)}')
        
        for f in sheet_formulas[:10]:
            print(f'    [{f["cell"]}] {f["formula"][:80]}')
        if len(sheet_formulas) > 10:
            print(f'    ... e altre {len(sheet_formulas) - 10} formule')

# 3. Dipendenze tra fogli
print('\n## DIPENDENZE TRA FOGLI (riferimenti cross-sheet):')
if cross_sheet_refs:
    for ref in cross_sheet_refs[:20]:
        print(f'  [{ref["sheet"]}!{ref["cell"]}] -> {ref["formula"][:60]}')
    if len(cross_sheet_refs) > 20:
        print(f'  ... e altri {len(cross_sheet_refs) - 20} riferimenti')
else:
    print('  Nessun riferimento cross-sheet trovato.')

# 4. Analisi struttura dati per import listino
print('\n## STRUTTURA DATI PER IMPORT LISTINO:')
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n### Foglio: {sheet_name}')
    
    headers = []
    for col in range(1, min(ws.max_column + 1, 20)):
        val = ws.cell(row=1, column=col).value
        if val:
            headers.append(f'{get_column_letter(col)}: {val}')
    
    print('  Intestazioni: ' + ', '.join(headers[:15]))
    
    print('  Prime righe dati:')
    for row in range(2, min(6, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(ws.max_column + 1, 10)):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                row_data.append(str(val)[:25])
        if row_data:
            print(f'    Riga {row}: {" | ".join(row_data)}')

# 5. Analisi formule critiche
print('\n## ANALISI FORMULE CRITICHE:')
critical_patterns = {
    'SOMMA/SUM': r'SUM\(|SOMMA\(',
    'CERCA/LOOKUP': r'LOOKUP|CERCA|VLOOKUP|HLOOKUP|INDEX|MATCH',
    'IF/SE': r'IF\(|SE\(',
    'Riferimento $': r'\$[A-Z]+\$?\d+|\$?\d+\$[A-Z]+',
    'Moltiplicazione': r'\*',
    'Divisione': r'/',
    'Concatenazione': r'&|CONCAT'
}

for sheet_name, formulas in formula_cells.items():
    print(f'\n### {sheet_name}:')
    for pattern_name, pattern in critical_patterns.items():
        matches = [f for f in formulas if re.search(pattern, f['formula'], re.IGNORECASE)]
        if matches:
            print(f'  {pattern_name}: {len(matches)} occorrenze')
            for m in matches[:3]:
                print(f'    [{m["cell"]}] {m["formula"][:60]}')

# 6. Cerca formule potenzialmente ridondanti
print('\n## FORMULE POTENZIALMENTE RIDONDANTI:')
formula_texts = {}
for sheet_name, formulas in formula_cells.items():
    for f in formulas:
        # Normalizza formula per trovare duplicati
        normalized = re.sub(r'\d+', 'N', f['formula'])
        if normalized not in formula_texts:
            formula_texts[normalized] = []
        formula_texts[normalized].append(f'{sheet_name}!{f["cell"]}')

duplicates = {k: v for k, v in formula_texts.items() if len(v) > 3}
if duplicates:
    for pattern, cells in list(duplicates.items())[:10]:
        print(f'  Pattern: {pattern[:50]}')
        print(f'    Trovato in {len(cells)} celle: {", ".join(cells[:5])}...')
else:
    print('  Nessuna formula significativamente ridondante.')

# 7. Verifica riferimenti circolari (semplificata)
print('\n## POTENZIALI RIFERIMENTI CIRCOLARI:')
print('  (Analisi basica - verificare manualmente)')
for sheet_name, formulas in formula_cells.items():
    for f in formulas:
        cell = f['cell']
        formula = f['formula']
        # Cerca se la cella si riferisce a se stessa
        if cell in formula:
            print(f'  ATTENZIONE: {sheet_name}!{cell} potrebbe riferirsi a se stessa: {formula[:50]}')

wb.close()
print('\n' + '=' * 80)
print('ANALISI COMPLETATA')
print('=' * 80)
